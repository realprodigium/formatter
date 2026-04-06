import openpyxl

wb = openpyxl.load_workbook(r'data.xlsx')
ws = wb['data']
lines = []
lines.append(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
for i, row in enumerate(ws.iter_rows(values_only=True)):
    vals = [str(v)[:60] if v is not None else 'NULL' for v in row]
    lines.append(f'R{i+1:03d}: ' + ' | '.join(vals))

with open('inspect_out.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))
print('Done - see inspect_out.txt')
