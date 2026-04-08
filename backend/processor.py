import pandas as pd
import re
from io import BytesIO
from datetime import datetime

class BankProcessor:
    @staticmethod
    def clean_currency(value):
        if pd.isna(value) or value == 'NULL':
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        
        # Remove $, spaces, and dots (thousands separator in some locales, but here it looks like pure digits)
        # Based on example: "$ 254115530" or "-$ 97698"
        clean_str = str(value).replace('$', '').replace(' ', '').replace(',', '')
        try:
            return float(clean_str)
        except ValueError:
            return 0.0

    @staticmethod
    def normalize_str(s):
        import unicodedata
        if not isinstance(s, str): return str(s)
        return "".join(
            c for c in unicodedata.normalize('NFD', s.lower())
            if unicodedata.category(c) != 'Mn'
        ).strip()

    @staticmethod
    def process_excel(file_content: bytes) -> bytes:
        try:
            df = pd.read_excel(BytesIO(file_content))
        except Exception as e:
            return b""
            
        if df.empty:
            return b""

        # Clean currency columns specifically for 'VALOR' and 'SALDO'
        for col in df.columns:
            col_norm = BankProcessor.normalize_str(col)
            if 'valor' in col_norm or 'saldo' in col_norm:
                df[col] = df[col].apply(BankProcessor.clean_currency)

        # Create a processed output
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
            
            # Access openpyxl objects to create a Table
            workbook = writer.book
            worksheet = writer.sheets['Data']
            
            from openpyxl.worksheet.table import Table, TableStyleInfo
            
            # Define table range (e.g., A1:H204)
            max_row = len(df) + 1
            max_col = len(df.columns)
            from openpyxl.utils import get_column_letter
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            
            tab = Table(displayName="TablaProcesada", ref=ref)
            
            # Add a default style
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                  showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            worksheet.add_table(tab)
            
            # Auto-adjust columns width
            for i, col in enumerate(df.columns, 1):
                column_letter = get_column_letter(i)
                # Compute max length of string representation of values in this column
                # We use a list comprehension to avoid pandas map(len) issues with non-strings
                val_lengths = [len(str(val)) for val in df[col].values if val is not None]
                col_name_len = len(str(col))
                
                max_len = max(val_lengths) if val_lengths else 0
                max_len = max(max_len, col_name_len) + 2
                
                worksheet.column_dimensions[column_letter].width = min(max_len, 50)

        return output.getvalue()
